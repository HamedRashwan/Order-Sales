from rest_framework import generics
from rest_framework.permissions import IsAuthenticated

from .models import Product, Customer, SalesOrder, StockMovement
from .serializers import (
    ProductSerializer,
    CustomerSerializer,
    SalesOrderSerializer,
    StockMovementSerializer,
)
from .permissions import (
    ProductPermission,
    CustomerPermission,
    SalesOrderPermission,
)
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework import generics, permissions
from django.contrib.auth.models import User

from .serializers import (
    ProductSerializer,
    CustomerSerializer,
    SalesOrderSerializer,
    StockMovementSerializer,
    UserRegisterSerializer,
)


class UserRegisterAPIView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserRegisterSerializer
    permission_classes = [permissions.AllowAny]

class ProductListAPIView(generics.ListAPIView):
    queryset = Product.objects.all().order_by("id")
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated, ProductPermission]


class ProductCreateAPIView(generics.CreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated, ProductPermission]


class ProductRetrieveAPIView(generics.RetrieveAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated, ProductPermission]


class ProductUpdateAPIView(generics.UpdateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated, ProductPermission]


class ProductDeleteAPIView(generics.DestroyAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated, ProductPermission]

# ========= CUSTOMERS =========

class CustomerListAPIView(generics.ListAPIView):
    queryset = Customer.objects.all().order_by("id")
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated, CustomerPermission]


class CustomerCreateAPIView(generics.CreateAPIView):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated, CustomerPermission]


class CustomerRetrieveAPIView(generics.RetrieveAPIView):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated, CustomerPermission]


class CustomerUpdateAPIView(generics.UpdateAPIView):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated, CustomerPermission]


class CustomerDeleteAPIView(generics.DestroyAPIView):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated, CustomerPermission]
# ========= SALES ORDERS =========

class SalesOrderListAPIView(generics.ListAPIView):
    queryset = (
        SalesOrder.objects.all()
        .select_related("customer", "created_by")
        .prefetch_related("items")
    )
    serializer_class = SalesOrderSerializer
    permission_classes = [IsAuthenticated, SalesOrderPermission]


class SalesOrderCreateAPIView(generics.CreateAPIView):
    queryset = SalesOrder.objects.all()
    serializer_class = SalesOrderSerializer
    permission_classes = [IsAuthenticated, SalesOrderPermission]


class SalesOrderRetrieveAPIView(generics.RetrieveAPIView):
    queryset = SalesOrder.objects.all().select_related("customer", "created_by")
    serializer_class = SalesOrderSerializer
    permission_classes = [IsAuthenticated, SalesOrderPermission]


class SalesOrderUpdateAPIView(generics.UpdateAPIView):
    queryset = SalesOrder.objects.all().select_related("customer", "created_by")
    serializer_class = SalesOrderSerializer
    permission_classes = [IsAuthenticated, SalesOrderPermission]


class SalesOrderDeleteAPIView(generics.DestroyAPIView):
    queryset = SalesOrder.objects.all()
    serializer_class = SalesOrderSerializer
    permission_classes = [IsAuthenticated, SalesOrderPermission]

# ========= STOCK MOVEMENTS =========

class StockMovementListAPIView(generics.ListAPIView):
    queryset = (
        StockMovement.objects.all()
        .select_related("product", "user")
        .order_by("-timestamp")
    )
    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated]


class StockMovementRetrieveAPIView(generics.RetrieveAPIView):
    queryset = StockMovement.objects.all().select_related("product", "user")
    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated]

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

class ProductsExcelReportAPIView(generics.ListAPIView):
    queryset = Product.objects.all().order_by("id")
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated, ProductPermission]

    def get(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())  # ✅ نفس filters/search/order لو موجودين

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = ["ID", "SKU", "Name", "Category", "Cost", "Selling", "Stock"]
        ws.append(headers)

        for p in qs:
            ws.append([p.id, p.sku, p.name, p.category, float(p.cost_price), float(p.selling_price), p.stock_qty])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="products_report.xlsx"'
        wb.save(resp)
        return resp
